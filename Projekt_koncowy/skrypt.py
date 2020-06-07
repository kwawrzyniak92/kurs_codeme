import os
import sys
from jinja2 import Environment, FileSystemLoader
import platform
import openpyxl
import unicodedata


class bcolors:
    ENDC = '\033[0m'
    BOLD = '\033[1m'


def clear_screen():
    if platform.system() == "Windows":
        os.system("cls")
    else:
        os.system("clear")


# Load the enviroment
THIS_DIR = os.path.dirname(os.path.abspath(__file__))

TEMPLATE_DIR = THIS_DIR + "\Templates"

env = Environment(loader=FileSystemLoader(TEMPLATE_DIR))

def sas():
    template = env.get_template('LTD_SAS.j2')


    class bcolors:
        HEADER = '\033[95m'
        OKBLUE = '\033[94m'
        OKGREEN = '\033[92m'
        WARNING = '\033[93m'
        FAIL = '\033[91m'
        ENDC = '\033[0m'
        BOLD = '\033[1m'
        UNDERLINE = '\033[4m'
        RED = '\33[41m'

    def z_pliku():

        XLSX_FILE_PATH = "C:\\Users\\Kamil\\Desktop\\Kurs_Code\\venv\\Projekt_koncowy\\clients_db.xlsx"

        book = openpyxl.load_workbook(XLSX_FILE_PATH, read_only=True)
        sheet = book.active

        num_rows = sheet.max_row
        num_cols = sheet.max_column


        LTD = {}

        i = eval(input("Podaj numer wiersza w pliku: "))

        if i in range(1, num_rows + 1):
            if sheet.cell(row=i, column=3).value == "LTD":
                if sheet.cell(row=i, column=2).value == "FO":
                    for j in range(1, num_cols + 1):
                        print(sheet.cell(row=i, column=j).value)
                        if j == 9:

                            LTD["customer"] = sheet.cell(row=i, column=j).value

                        elif j == 14:

                            LTD["contact"] = sheet.cell(row=i, column=j).value


                        elif j == 8:

                            LTD["customer_id"] = sheet.cell(row=i, column=j).value

                        elif j == 10:

                            LTD["service_id"] = sheet.cell(row=i, column=j).value

                        elif j == 13:

                            LTD["port_number"] = sheet.cell(row=i, column=j).value

                        elif j == 15:

                            LTD["bandwith"] = sheet.cell(row=i, column=j).value


                        elif j == 6:

                            LTD["sdp_id_primary"] = sheet.cell(row=i, column=j).value

                        elif j == 7:

                            LTD["sdp_id_secondary"] = sheet.cell(row=i, column=j).value
                        elif j == 12:

                            LTD["switch_name"] = sheet.cell(row=i, column=j).value
                        elif j == 11:

                            LTD["vlan"] = sheet.cell(row=i, column=j).value

                   # pobieranie danych o nazwie klienta - dzielone znakiem ;
                    customer_name = LTD["customer"].split(';')[0]
                    if customer_name[0] == ' ':
                        customer_name = customer_name[1:]
                    customer_name = customer_name.replace(' ', '-')
                    customer_name = unicodedata.normalize("NFD", customer_name)
                    customer_name = customer_name.replace('\u0142', 'l')
                    customer_name = customer_name.replace('\u0141', 'L')
                    customer_name = customer_name.replace('\u0301','')


                    customer_adress = LTD["customer"].split(';')[1]
                    if customer_adress[0] == ' ':
                        customer_adress = customer_adress[1:]
                    if customer_adress.endswith(' - ltd'):
                        customer_adress = customer_adress[:-6]
                    if customer_adress.endswith('-ltd'):
                        customer_adress = customer_adress[:-4]
                    if customer_adress.endswith(' -ltd'):
                        customer_adress = customer_adress[:-5]
                    if customer_adress.endswith('- ltd'):
                        customer_adress = customer_adress[:-5]
                    customer_adress = customer_adress.replace(',', '')
                    customer_adress = customer_adress.replace(' ', '-')
                    customer_adress = unicodedata.normalize("NFD", customer_adress)
                    customer_adress = customer_adress.replace('\u0142', 'l')
                    customer_adress = customer_adress.replace('\u0141', 'L')
                    customer_adress = customer_adress.replace('\u0301','')

                    LTD["customer_name"] = str(customer_name)
                    LTD["customer_adress"] = str(customer_adress)

                    # wyciaganie danych kontaktowych
                    contact = LTD["contact"].replace(' ', '-', 1)
                    contact = unicodedata.normalize("NFD", contact)
                    contact = contact.replace('\u0142', 'l')
                    contact = contact.replace('\u0141', 'L')
                    contact = contact.replace('\u0301','')
                    # remove_digits = str.maketrans('','',digits)
                    # LTD["customer_contact"] = contact.translate(remove_digits)
                    LTD["customer_contact"] = contact
                    # LTD["customer_phone"] = int(list(filter(str.isdigit, contact)))
                    LTD["customer_phone"] = "666222333"

                    # przeliczanie pasma
                    a, b = LTD["bandwith"].split('/')
                    LTD["down_rate"] = int(a) * 1000
                    LTD["up_rate"] = int(b) * 1050

                    # zamiana portu na numer qos
                    qos = LTD["port_number"].split("/")[2]
                    LTD["qos"] = str(qos + '0')
                    #generowanie konfiguracji z templatea i wyświetlanie użytkownikowi
                    output = template.render(LTD)
                    print(output)

                    #zapisywanie pliku w folderze Services
                    filename = LTD["customer_name"] + "_" + LTD["customer_adress"] + "_" + "LTD-DC.txt"
                    #!!!!!!!!!!!!!!!!!!ŚCIEŻKA ZAPISU PLIKU!!!!!!!!
                    filepath = "C:\\Users\\Kamil\\Desktop\\Kurs_Code\\venv\\Projekt_koncowy\\Services\\" + filename

                    file = open(filepath, "w")
                    file.write(output)
                    file.close()
                    book.close()

                else:
                    print(bcolors.RED + "\n TO NIE JEST FO\n\n" + bcolors.ENDC)
                    menu()
            else:
                print(bcolors.RED + "\n TO NIE JEST USŁUGA LTD\n\n" + bcolors.ENDC)
                menu()

    print(bcolors.BOLD + "-------------------------------------------------------------" + bcolors.ENDC)
    print(bcolors.BOLD + "------------ Generowanie konfiguracji połączenia ------------" + bcolors.ENDC)
    print(bcolors.BOLD + "--------------------- Klinent - DC --------------------------" + bcolors.ENDC)
    print(bcolors.BOLD + "-------------------------------------------------------------\n" + bcolors.ENDC)

    input_choice = input(bcolors.BOLD + "===== PODAJ DANE USŁUGI =====\n"
                                        "1. Pobierz z pliku clients_db\n"
                                        "2. Wpisz ręcznie\n"
                                        "3. Wyjście"
                                        "\n\nWprowadź swój wybór [1-3] i wciśnij [Enter]: " + bcolors.ENDC)

    while (input_choice == "1") or (input_choice == "2") or (input_choice == "3"):

        if input_choice == "1":
            clear_screen()
            z_pliku()
            break
        elif input_choice == "2":
            clear_screen()
            print("w trakcie przygotowania")
            # manual_insert()
            break
        elif input_choice == "3":
            sys.exit()


def menu():
    choice = input(bcolors.BOLD + "=======================================================\n"
                                  "===== Do jakiego urządzenia podłączony jest klient =====\n"
                                  "========================================================\n\n"
                                  "1. SAS\n"
                                  "2. A8\n"
                                  "3. Wyjście"
                                  "\n\nWprowadź swój wybór [1-3] i wciśnij [Enter]: " + bcolors.ENDC)

    while (choice == "1") or (choice == "2") or (choice == "3"):

        if choice == "1":
            clear_screen()
            sas()
            break
        elif choice == "2":
            clear_screen()
            print("w trakcie przygotowania")
            # a8()
            break
        elif choice == "3":
            sys.exit()


def main():
    clear_screen()
    menu()

    main_choice = input("\nWcisnij '1' by zaczac od nowa lub [Enter] aby zakończyć: ")

    while main_choice == "1":
        clear_screen()
        menu()

if __name__ == "__main__":
    main()
